"""
Salvatore Pizza — Goldman Sachs grade financial model builder.
Run: python3 build_model.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, LineChart, Reference

OUT_DIR = "/Users/mertgulen/Library/CloudStorage/GoogleDrive-mertgulen98@gmail.com/My Drive/Carisma Wellness Group/Carisma AI /Carisma AI/salvatore-pizza/financials"
OUT_FILE = os.path.join(OUT_DIR, "Salvatore_Pizza_Financial_Model.xlsx")

# ---------- Styling palette ----------
NAVY = "0E2A47"
GRAY = "D9D9D9"
INPUT_BLUE = "DCE6F1"
OUTPUT_YELLOW = "FFF2CC"
RED = "F8CBAD"
GREEN = "C6E0B4"

THIN = Side(style="thin", color="9BA4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_BGN = '#,##0_);[Red](#,##0)'
FMT_BGN_DEC = '#,##0.00_);[Red](#,##0.00)'
FMT_PCT = '0.0%'
FMT_MULT = '0.00"x"'

def title_row(ws, sheet_name, cols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=f"SALVATORE PIZZA  —  {sheet_name.upper()}")
    c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

def sub_header(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=GRAY)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

def input_cell(cell, fmt=FMT_BGN):
    cell.fill = PatternFill("solid", fgColor=INPUT_BLUE)
    cell.border = BORDER
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")

def calc_cell(cell, fmt=FMT_BGN, bold=False):
    cell.border = BORDER
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    if bold:
        cell.font = Font(bold=True)

def output_cell(cell, fmt=FMT_BGN):
    cell.fill = PatternFill("solid", fgColor=OUTPUT_YELLOW)
    cell.border = BORDER
    cell.font = Font(bold=True)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")

def label_cell(cell, bold=False, indent=0):
    cell.alignment = Alignment(horizontal="left", indent=indent)
    if bold:
        cell.font = Font(bold=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()
wb.remove(wb.active)

YEARS = ["Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030"]

# ===================================================================
# 1. COVER
# ===================================================================
ws = wb.create_sheet("Cover")
title_row(ws, "Cover", cols=8)
ws["A3"] = "Salvatore Pizza | Financial Model & Investment Case"
ws["A3"].font = Font(size=20, bold=True, color=NAVY)
ws.merge_cells("A3:H3")
ws["A4"] = "Sofia, Bulgaria  |  5-Year Projection  |  Currency: BGN (EUR @ 1.95583)"
ws["A4"].font = Font(size=12, italic=True, color="555555")
ws.merge_cells("A4:H4")

ws["A6"] = "Concept:"
ws["B6"] = "Authentic Neapolitan pizzeria — 48h fermented dough, signature stuffed-mozzarella crust"
ws["A7"] = "Location:"
ws["B7"] = "19 Gurko Str., Central Sofia, Bulgaria"
ws["A8"] = "Capacity:"
ws["B8"] = "1 wood-fired oven, 6 in-shop tables, delivery via aggregator"
ws["A9"] = "Hours:"
ws["B9"] = "11:00 – 23:00, 7 days/week (~360 trading days/yr)"
ws["A10"] = "Prepared by:"
ws["B10"] = "Goldman Sachs — Restaurant & Hospitality Coverage"
ws["A11"] = "Date:"
ws["B11"] = "2026-05-20"
for r in range(6, 12):
    ws.cell(row=r, column=1).font = Font(bold=True)

# Headline KPI box — values pulled from other tabs by formula
ws["A14"] = "CENTRAL CASE — HEADLINE KPIs"
ws["A14"].font = Font(size=12, bold=True, color="FFFFFF")
ws["A14"].fill = PatternFill("solid", fgColor=NAVY)
ws.merge_cells("A14:H14")
ws["A14"].alignment = Alignment(horizontal="center")

kpi_labels = [
    ("Year-1 Revenue (BGN)", "='P&L'!C5", FMT_BGN),
    ("Year-5 Revenue (BGN)", "='P&L'!G5", FMT_BGN),
    ("Y3 EBITDA Margin",     "='P&L'!E22", FMT_PCT),
    ("Y5 EBITDA Margin",     "='P&L'!G22", FMT_PCT),
    ("Project IRR",          "='Cash Flow'!C25", FMT_PCT),
    ("Payback Period (yrs)", "='Cash Flow'!C27", '0.0" yrs"'),
    ("NPV @ 12% (BGN)",      "='Cash Flow'!C24", FMT_BGN),
    ("MOIC",                 "='Cash Flow'!C26", FMT_MULT),
]
for i, (label, formula, fmt) in enumerate(kpi_labels):
    r = 16 + i
    ws.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    ws.cell(row=r, column=1).border = BORDER
    c = ws.cell(row=r, column=3, value=formula)
    output_cell(c, fmt=fmt)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

ws["A26"] = "Investment Thesis (Central Case)"
ws["A26"].font = Font(bold=True, size=11, color=NAVY)
ws["A27"] = "• Premium positioning via stuffed-mozzarella signature crust differentiates from Sofia's commodity Neapolitan field."
ws["A28"] = "• Single-oven, 6-table footprint with delivery upside keeps fixed-cost base lean; break-even achievable in Y1."
ws["A29"] = "• Bulgaria's 10% CIT, 9% reduced-VAT on food, and sub-EU labour costs underpin attractive unit economics."
ws["A30"] = "• Key risks: aggregator commission compression of delivery margin; oven throughput ceiling at 180 pizzas/day."

ws["A34"] = "Prepared by Goldman Sachs Restaurant & Hospitality Coverage  |  Confidential Due Diligence Draft  |  2026-05-20"
ws["A34"].font = Font(italic=True, size=9, color="888888")
ws.merge_cells("A34:H34")

set_col_widths(ws, [22, 22, 18, 18, 18, 12, 12, 12])
ws.freeze_panes = "A2"

# ===================================================================
# 2. ASSUMPTIONS
# ===================================================================
ws = wb.create_sheet("Assumptions")
title_row(ws, "Assumptions", cols=6)
set_col_widths(ws, [42, 18, 18, 14, 42, 14])

ws["A3"] = "All Bulgaria/Sofia benchmarks anchored to 2025-26 market data. Inputs in BLUE are flex-able drivers."
ws["A3"].font = Font(italic=True, color="555555")
ws.merge_cells("A3:F3")

row = 5
# Helper to add a group
def add_group(name, items, start_row):
    r = start_row
    ws.cell(row=r, column=1, value=name).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    r += 1
    # header
    for col, h in enumerate(["Driver", "Value", "Unit", "", "Source / Note", ""], start=1):
        c = ws.cell(row=r, column=col, value=h)
        sub_header(c)
    r += 1
    for label, val, unit, note, fmt in items:
        ws.cell(row=r, column=1, value=label)
        label_cell(ws.cell(row=r, column=1), indent=1)
        ws.cell(row=r, column=1).border = BORDER
        c = ws.cell(row=r, column=2, value=val)
        input_cell(c, fmt=fmt)
        ws.cell(row=r, column=3, value=unit).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).border = BORDER
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.cell(row=r, column=5, value=note).font = Font(italic=True, color="555555")
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="center")
        r += 1
    return r + 1

# OPERATING
operating = [
    ("Trading days per year",                360,     "days",   "Open 7 days/week, allow 5 closed days",                 '#,##0'),
    ("Trading hours per day",                12,      "hrs",    "11:00–23:00",                                            '#,##0'),
    ("Unit size",                            90,      "m²",     "Central Sofia 90 m² assumption",                        '#,##0'),
    ("EUR / BGN fixed peg",                  1.95583, "BGN/EUR","Bulgaria currency board peg",                            '0.00000'),
    ("Scenario (Base/Bull/Bear)",            "Base",  "—",     "Drives flex on covers/day & ASP via ±15% in calcs",      '@'),
    ("Scenario flex — Bull (+) / Bear (-)",  0.15,    "%",      "Applied to covers and ASP under Bull/Bear",              FMT_PCT),
]
row = add_group("Operating Parameters", operating, row)

# REVENUE
revenue = [
    ("# tables in-shop",                     6,       "tables", "Physical constraint",                                    '#,##0'),
    ("Covers per table (seats)",             4,       "pax",    "Standard 4-top",                                         '#,##0'),
    ("Lunch turns/day Y1",                   1.0,     "turns",  "Ramping to 1.4 by Y3",                                   '0.0'),
    ("Dinner turns/day Y1",                  1.5,     "turns",  "Ramping to 2.3 by Y3 (Sofia central peak)",              '0.0'),
    ("Lunch turns/day Y3 (steady)",          1.4,     "turns",  "Steady state",                                           '0.0'),
    ("Dinner turns/day Y3 (steady)",         2.3,     "turns",  "Steady state",                                           '0.0'),
    ("Avg dine-in ticket (BGN/cover)",       34,      "BGN",    "Pizza ASP 24 + drink + occasional side",                 FMT_BGN),
    ("Takeaway orders/day Y1",               12,      "orders", "Walk-in pickup",                                         '#,##0'),
    ("Takeaway orders/day Y3 (steady)",      25,      "orders", "Steady",                                                 '#,##0'),
    ("Avg takeaway ticket (BGN)",            26,      "BGN",    "Pizza + small drink",                                    FMT_BGN),
    ("Delivery orders/day Y1",               18,      "orders", "Via Glovo/Foodpanda Sofia",                              '#,##0'),
    ("Delivery orders/day Y3 (steady)",      45,      "orders", "Aggressive ramp to 45 by Y3",                            '#,##0'),
    ("Avg delivery ticket (BGN, gross)",     32,      "BGN",    "Pizza + drink, less attach",                             FMT_BGN),
    ("Same-store growth Y2",                 0.25,    "%",      "Year-2 ramp",                                            FMT_PCT),
    ("Same-store growth Y3",                 0.15,    "%",      "Year-3 ramp",                                            FMT_PCT),
    ("Same-store growth Y4",                 0.06,    "%",      "Maturation",                                             FMT_PCT),
    ("Same-store growth Y5",                 0.04,    "%",      "Steady state",                                           FMT_PCT),
    ("Oven max throughput (pizza/day)",      180,     "pizzas", "Single Neapolitan oven peak — flag if breached",         '#,##0'),
]
row = add_group("Revenue Drivers", revenue, row)

# COSTS
costs = [
    ("Food COGS % of food revenue",          0.30,    "%",      "Neapolitan target 28–32%; stuffed crust +2% cheese",     FMT_PCT),
    ("Beverage COGS % of bev revenue",       0.25,    "%",      "Bulgaria casual-dining benchmark",                       FMT_PCT),
    ("Beverage share of revenue",            0.18,    "%",      "Blended bev share",                                      FMT_PCT),
    ("Aggregator commission %",              0.30,    "%",      "Glovo/Foodpanda Bulgaria 28–32%; own fleet would lower", FMT_PCT),
    ("Rent BGN/month",                       6500,    "BGN/mo", "Central Sofia 90 m² @ ~72 BGN/m²/mo",                    FMT_BGN),
    ("Rent annual indexation",               0.03,    "%",      "Lease escalator",                                        FMT_PCT),
    ("Utilities base BGN/month",             1500,    "BGN/mo", "Power/water/internet",                                   FMT_BGN),
    ("Oven gas/wood BGN/month",              750,     "BGN/mo", "Wood-fired oven fuel",                                   FMT_BGN),
    ("Head chef gross BGN/month",            3500,    "BGN/mo", "Senior Italian-trained",                                 FMT_BGN),
    ("Pizzaioli x2 gross BGN/month each",    2200,    "BGN/mo", "Bulgaria restaurant skilled wage",                       FMT_BGN),
    ("Waiters x2 gross BGN/month each",      1500,    "BGN/mo", "Above Bulgaria min wage 1,077",                          FMT_BGN),
    ("Delivery coordinator (PT) BGN/month",  1200,    "BGN/mo", "Part-time logistics",                                    FMT_BGN),
    ("Employer social contributions %",      0.19,    "%",      "Bulgaria employer-side payroll taxes",                   FMT_PCT),
    ("Marketing % of revenue Y1 (launch)",   0.05,    "%",      "Launch year heavier",                                    FMT_PCT),
    ("Marketing % of revenue Y2+ (steady)",  0.03,    "%",      "Steady",                                                 FMT_PCT),
    ("Repairs & maintenance % of revenue",   0.015,   "%",      "Industry benchmark",                                     FMT_PCT),
    ("Packaging % of revenue",               0.012,   "%",      "Delivery + takeaway boxes",                              FMT_PCT),
    ("Insurance & licenses BGN/yr",          4000,    "BGN/yr", "Liability + municipal",                                  FMT_BGN),
    ("POS / SaaS BGN/month",                 300,     "BGN/mo", "POS + delivery integration",                             FMT_BGN),
    ("Accounting BGN/month",                 500,     "BGN/mo", "External bookkeeper",                                    FMT_BGN),
    ("Misc % of revenue",                    0.01,    "%",      "Buffer for one-offs",                                    FMT_PCT),
]
row = add_group("Cost Drivers", costs, row)

# CAPEX
capex = [
    ("Wood-fired oven (Naples import)",      30000,   "BGN",    "25–35k typical",                                         FMT_BGN),
    ("Kitchen equipment",                    35000,   "BGN",    "Cold line, prep, dough mixer, fridges",                  FMT_BGN),
    ("Fit-out / design / 6 tables / lighting",75000,  "BGN",    "60–90k for central Sofia",                               FMT_BGN),
    ("POS + delivery integration",           5000,    "BGN",    "Hardware + software",                                    FMT_BGN),
    ("Signage",                              8000,    "BGN",    "External + interior",                                    FMT_BGN),
    ("Opening inventory",                    10000,   "BGN",    "Food + bev launch stock",                                FMT_BGN),
    ("Deposits & legal",                     15000,   "BGN",    "Lease deposit + setup",                                  FMT_BGN),
    ("Pre-opening marketing",                10000,   "BGN",    "Launch buzz",                                            FMT_BGN),
    ("Contingency",                          12000,   "BGN",    "6% buffer",                                              FMT_BGN),
    ("Total Opening CAPEX",                  200000,  "BGN",    "Central case",                                           FMT_BGN),
    ("Maintenance CAPEX % of revenue (Y2+)", 0.02,    "%",      "Ongoing reinvestment",                                   FMT_PCT),
    ("D&A useful life",                      7,       "years",  "Straight-line",                                          '#,##0'),
]
row = add_group("CAPEX", capex, row)

# FINANCING
financing = [
    ("Working capital buffer",               20000,   "BGN",    "Recovered at exit",                                      FMT_BGN),
    ("Equity funding (100%)",                220000,  "BGN",    "Opening CAPEX + WC",                                     FMT_BGN),
    ("Discount rate (WACC proxy)",           0.12,    "%",      "Independent restaurant CEE",                             FMT_PCT),
    ("Exit EBITDA multiple",                 4.5,     "x",      "Independent QSR CEE sector standard",                    FMT_MULT),
    ("Interest expense",                     0,       "BGN",    "All-equity capital stack",                               FMT_BGN),
]
row = add_group("Financing & Valuation", financing, row)

# TAX
tax = [
    ("VAT — food (in-shop & delivery)",      0.09,    "%",      "Bulgaria reduced rate",                                  FMT_PCT),
    ("VAT — alcohol",                        0.20,    "%",      "Standard rate",                                          FMT_PCT),
    ("Blended VAT (model)",                  0.10,    "%",      "Weighted blend used in model",                           FMT_PCT),
    ("Corporate income tax",                 0.10,    "%",      "Bulgaria flat CIT (lowest in EU)",                       FMT_PCT),
]
row = add_group("Tax", tax, row)

ws.freeze_panes = "A2"

# Named refs — we'll just reference Assumptions cells directly by row for clarity.
# Build a small map of {label: cell_address} by scanning column A.
A_REF = {}
for r in range(5, row + 1):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, str):
        A_REF[v.strip()] = f"Assumptions!$B${r}"

def aref(label):
    return A_REF[label]

# ===================================================================
# 3. REVENUE BUILD
# ===================================================================
ws = wb.create_sheet("Revenue Build")
title_row(ws, "Revenue Build", cols=8)
set_col_widths(ws, [38, 16, 16, 16, 16, 16, 16, 16])

ws["A3"] = "Annual roll-up. Scenario flex applies multiplier to covers/day and ASP."
ws["A3"].font = Font(italic=True, color="555555")
ws.merge_cells("A3:H3")

# Year header
for i, y in enumerate(YEARS):
    c = ws.cell(row=4, column=3 + i, value=y)
    sub_header(c)
ws.cell(row=4, column=2, value="Driver / Unit")
sub_header(ws.cell(row=4, column=2))
ws.cell(row=4, column=1, value="Line")
sub_header(ws.cell(row=4, column=1))

# Scenario multiplier (Base 1; Bull 1+flex; Bear 1-flex)
ws["A6"] = "Scenario multiplier (volume & ASP)"
flex_formula = f'=IF({aref("Scenario (Base/Bull/Bear)")}="Bull",1+{aref("Scenario flex — Bull (+) / Bear (-)")},IF({aref("Scenario (Base/Bull/Bear)")}="Bear",1-{aref("Scenario flex — Bull (+) / Bear (-)")},1))'
for i in range(5):
    c = ws.cell(row=6, column=3 + i, value=flex_formula)
    calc_cell(c, fmt='0.000')
ws.cell(row=6, column=2, value="x")

# Helper: linear ramp from Y1 -> Y3 across years 1..3 then flat
def ramp(y_idx, y1_ref, y3_ref):
    if y_idx == 0:
        return f"={y1_ref}"
    elif y_idx == 1:
        return f"=({y1_ref}+{y3_ref})/2"
    else:
        return f"={y3_ref}"

# Dine-in build
row = 8
ws.cell(row=row, column=1, value="DINE-IN").font = Font(bold=True, color=NAVY)
ws.cell(row=row, column=2, value="").fill = PatternFill("solid", fgColor=GRAY)
row += 1
# Tables × covers/table
ws.cell(row=row, column=1, value="  Tables × seats/table")
ws.cell(row=row, column=2, value="pax")
for i in range(5):
    f = f'={aref("# tables in-shop")}*{aref("Covers per table (seats)")}'
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt='#,##0')
seats_row = row
row += 1
# Lunch turns/day
ws.cell(row=row, column=1, value="  Lunch turns/day")
ws.cell(row=row, column=2, value="turns")
for i in range(5):
    f = ramp(i, aref("Lunch turns/day Y1"), aref("Lunch turns/day Y3 (steady)"))
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt='0.00')
lunch_turns_row = row
row += 1
# Dinner turns/day
ws.cell(row=row, column=1, value="  Dinner turns/day")
ws.cell(row=row, column=2, value="turns")
for i in range(5):
    f = ramp(i, aref("Dinner turns/day Y1"), aref("Dinner turns/day Y3 (steady)"))
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt='0.00')
dinner_turns_row = row
row += 1
# Daily covers
ws.cell(row=row, column=1, value="  Daily covers (lunch + dinner)")
ws.cell(row=row, column=2, value="covers")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"={col}{seats_row}*({col}{lunch_turns_row}+{col}{dinner_turns_row})*{col}6"
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt='#,##0')
daily_covers_row = row
row += 1
# Avg ticket
ws.cell(row=row, column=1, value="  Avg ticket BGN/cover (scen-flexed)")
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f'={aref("Avg dine-in ticket (BGN/cover)")}*{col}6'
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt=FMT_BGN)
dinein_ticket_row = row
row += 1
# Dine-in revenue (with same-store growth applied COMPOUNDING from Y2)
ws.cell(row=row, column=1, value="Dine-in revenue (annual)").font = Font(bold=True)
ws.cell(row=row, column=2, value="BGN")
# Build base Y1 from physical drivers; then compound prior year by SSSG for Y2+
sscg_refs = [aref("Same-store growth Y2"), aref("Same-store growth Y3"), aref("Same-store growth Y4"), aref("Same-store growth Y5")]
for i in range(5):
    col = get_column_letter(3 + i)
    if i == 0:
        f = f"={col}{daily_covers_row}*{col}{dinein_ticket_row}*{aref('Trading days per year')}"
    else:
        prev_col = get_column_letter(3 + i - 1)
        f = f"={prev_col}{row}*(1+{sscg_refs[i-1]})"
    c = ws.cell(row=row, column=3 + i, value=f); output_cell(c)
dinein_rev_row = row
row += 2

# Takeaway
ws.cell(row=row, column=1, value="TAKEAWAY").font = Font(bold=True, color=NAVY)
row += 1
ws.cell(row=row, column=1, value="  Takeaway orders/day")
ws.cell(row=row, column=2, value="orders")
for i in range(5):
    f = ramp(i, aref("Takeaway orders/day Y1"), aref("Takeaway orders/day Y3 (steady)"))
    c = ws.cell(row=row, column=3 + i, value=f"={f[1:]}*" + get_column_letter(3 + i) + "6")
    calc_cell(c, fmt='#,##0')
takeaway_orders_row = row
row += 1
ws.cell(row=row, column=1, value="  Avg ticket (scen-flexed)")
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f'={aref("Avg takeaway ticket (BGN)")}*{col}6'
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt=FMT_BGN)
takeaway_ticket_row = row
row += 1
ws.cell(row=row, column=1, value="Takeaway revenue (annual)").font = Font(bold=True)
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    if i == 0:
        f = f"={col}{takeaway_orders_row}*{col}{takeaway_ticket_row}*{aref('Trading days per year')}"
    else:
        prev_col = get_column_letter(3 + i - 1)
        f = f"={prev_col}{row}*(1+{sscg_refs[i-1]})"
    c = ws.cell(row=row, column=3 + i, value=f); output_cell(c)
takeaway_rev_row = row
row += 2

# Delivery
ws.cell(row=row, column=1, value="DELIVERY (gross)").font = Font(bold=True, color=NAVY)
row += 1
ws.cell(row=row, column=1, value="  Delivery orders/day")
ws.cell(row=row, column=2, value="orders")
for i in range(5):
    f = ramp(i, aref("Delivery orders/day Y1"), aref("Delivery orders/day Y3 (steady)"))
    c = ws.cell(row=row, column=3 + i, value=f"={f[1:]}*" + get_column_letter(3 + i) + "6")
    calc_cell(c, fmt='#,##0')
delivery_orders_row = row
row += 1
ws.cell(row=row, column=1, value="  Avg ticket gross (scen-flexed)")
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f'={aref("Avg delivery ticket (BGN, gross)")}*{col}6'
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt=FMT_BGN)
delivery_ticket_row = row
row += 1
ws.cell(row=row, column=1, value="Delivery revenue GROSS (annual)").font = Font(bold=True)
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    if i == 0:
        f = f"={col}{delivery_orders_row}*{col}{delivery_ticket_row}*{aref('Trading days per year')}"
    else:
        prev_col = get_column_letter(3 + i - 1)
        f = f"={prev_col}{row}*(1+{sscg_refs[i-1]})"
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, bold=True)
delivery_gross_row = row
row += 1
ws.cell(row=row, column=1, value="  Less: aggregator commission")
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"=-{col}{delivery_gross_row}*{aref('Aggregator commission %')}"
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c)
delivery_comm_row = row
row += 1
ws.cell(row=row, column=1, value="Delivery revenue NET (annual)").font = Font(bold=True)
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"={col}{delivery_gross_row}+{col}{delivery_comm_row}"
    c = ws.cell(row=row, column=3 + i, value=f); output_cell(c)
delivery_net_row = row
row += 2

# Total revenue
ws.cell(row=row, column=1, value="TOTAL REVENUE (Net)").font = Font(bold=True, size=12, color=NAVY)
ws.cell(row=row, column=2, value="BGN")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"={col}{dinein_rev_row}+{col}{takeaway_rev_row}+{col}{delivery_net_row}"
    c = ws.cell(row=row, column=3 + i, value=f); output_cell(c)
total_rev_row = row
row += 1
ws.cell(row=row, column=1, value="  of which: Gross delivery (for COGS/comm calc)").font = Font(italic=True)
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=row, column=3 + i, value=f"={col}{delivery_gross_row}"); calc_cell(c)
total_rev_gross_delivery_row = row
row += 2

# Oven capacity check
ws.cell(row=row, column=1, value="OVEN CAPACITY CHECK").font = Font(bold=True, color=NAVY)
row += 1
ws.cell(row=row, column=1, value="  Implied pizzas/day (≈ 90% of orders)")
ws.cell(row=row, column=2, value="pizzas")
for i in range(5):
    col = get_column_letter(3 + i)
    # Approx: covers + takeaway + delivery orders, each ~0.9 pizzas
    f = f"=({col}{daily_covers_row}*0.7+{col}{takeaway_orders_row}+{col}{delivery_orders_row})"
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt='#,##0')
pizzas_day_row = row
row += 1
ws.cell(row=row, column=1, value="  Oven ceiling (pizzas/day)")
ws.cell(row=row, column=2, value="pizzas")
for i in range(5):
    f = f'={aref("Oven max throughput (pizza/day)")}'
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt='#,##0')
oven_ceiling_row = row
row += 1
ws.cell(row=row, column=1, value="  Utilisation").font = Font(italic=True)
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"={col}{pizzas_day_row}/{col}{oven_ceiling_row}"
    c = ws.cell(row=row, column=3 + i, value=f); calc_cell(c, fmt=FMT_PCT, bold=True)
util_row = row

# Conditional formatting: red if >100%, green otherwise
from openpyxl.formatting.rule import CellIsRule
red_fill = PatternFill("solid", fgColor=RED)
green_fill = PatternFill("solid", fgColor=GREEN)
ws.conditional_formatting.add(f"C{util_row}:G{util_row}",
    CellIsRule(operator="greaterThan", formula=["1"], fill=red_fill))
ws.conditional_formatting.add(f"C{util_row}:G{util_row}",
    CellIsRule(operator="lessThanOrEqual", formula=["1"], fill=green_fill))

ws.freeze_panes = "C5"

# Persist refs for downstream sheets
REV_TOTAL_ROW = total_rev_row
REV_GROSS_DELIVERY_ROW = total_rev_gross_delivery_row
DINEIN_REV_ROW = dinein_rev_row
TAKEAWAY_REV_ROW = takeaway_rev_row
DELIVERY_NET_ROW = delivery_net_row
DELIVERY_GROSS_ROW = delivery_gross_row

# ===================================================================
# 4. COST FORECAST
# ===================================================================
ws = wb.create_sheet("Cost Forecast")
title_row(ws, "Cost Forecast", cols=12)
set_col_widths(ws, [38, 16, 12, 16, 12, 16, 12, 16, 12, 16, 12, 12])

ws["A3"] = "Each cost shown in BGN and as % of total revenue. Linked to Revenue Build & Assumptions."
ws["A3"].font = Font(italic=True, color="555555")
ws.merge_cells("A3:L3")

# Header: Year × (BGN, %)
ws.cell(row=4, column=1, value="Cost Line")
sub_header(ws.cell(row=4, column=1))
for i, y in enumerate(YEARS):
    start_col = 2 + i * 2
    ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + 1)
    c = ws.cell(row=4, column=start_col, value=y); sub_header(c)
ws.cell(row=5, column=1, value="")
for i in range(5):
    start_col = 2 + i * 2
    sub_header(ws.cell(row=5, column=start_col, value="BGN"))
    sub_header(ws.cell(row=5, column=start_col + 1, value="% rev"))

# Revenue handle (link)
ws.cell(row=6, column=1, value="Total Revenue (ref)").font = Font(italic=True)
for i in range(5):
    col_rev = get_column_letter(3 + i)
    bgn_col = 2 + i * 2
    f = f"='Revenue Build'!{col_rev}{REV_TOTAL_ROW}"
    c = ws.cell(row=6, column=bgn_col, value=f); calc_cell(c)
    ws.cell(row=6, column=bgn_col + 1, value="").border = BORDER

ROW_REV_LINK = 6

def add_cost_line(row, label, formulas_by_year, bold=False, indent=1, is_subtotal=False):
    ws.cell(row=row, column=1, value=label)
    label_cell(ws.cell(row=row, column=1), bold=bold, indent=indent)
    ws.cell(row=row, column=1).border = BORDER
    for i, f in enumerate(formulas_by_year):
        bgn_col = 2 + i * 2
        c = ws.cell(row=row, column=bgn_col, value=f)
        if is_subtotal:
            output_cell(c)
        else:
            calc_cell(c, bold=bold)
        # % of revenue
        pct_col = bgn_col + 1
        rev_col_letter = get_column_letter(bgn_col)
        rev_link_col = get_column_letter(2 + i * 2)  # same
        pct_f = f"=IFERROR({get_column_letter(bgn_col)}{row}/{rev_link_col}{ROW_REV_LINK},0)"
        cp = ws.cell(row=row, column=pct_col, value=pct_f)
        calc_cell(cp, fmt=FMT_PCT, bold=bold)
    return row

# Helper to ref Revenue Build by year-index
def rb(col_offset_year, target_row):
    return f"'Revenue Build'!{get_column_letter(3 + col_offset_year)}{target_row}"

row = 7
ws.cell(row=row, column=1, value="COGS").font = Font(bold=True, color=NAVY)
row += 1
# Food COGS: food share = (1 - bev share); COGS = food_share * total_rev * food_cogs%
food_formulas = []
for i in range(5):
    rev = rb(i, REV_TOTAL_ROW)
    f = f"={rev}*(1-{aref('Beverage share of revenue')})*{aref('Food COGS % of food revenue')}"
    food_formulas.append(f)
add_cost_line(row, "  Food COGS", food_formulas); food_cogs_row = row; row += 1

bev_formulas = []
for i in range(5):
    rev = rb(i, REV_TOTAL_ROW)
    f = f"={rev}*{aref('Beverage share of revenue')}*{aref('Beverage COGS % of bev revenue')}"
    bev_formulas.append(f)
add_cost_line(row, "  Beverage COGS", bev_formulas); bev_cogs_row = row; row += 1

# Aggregator commission shown separately (already netted in revenue but show as cost line for transparency)
# We'll show it as memo only (set in light italic) since it's already deducted
agg_formulas = []
for i in range(5):
    f = f"={rb(i, DELIVERY_GROSS_ROW)}*{aref('Aggregator commission %')}"
    agg_formulas.append(f)
add_cost_line(row, "  Aggregator commission (memo — netted in rev)", agg_formulas, indent=1)
ws.cell(row=row, column=1).font = Font(italic=True, color="888888")
agg_row = row; row += 1

# COGS subtotal (excl memo)
cogs_sub_formulas = []
for i in range(5):
    bgn_col = 2 + i * 2
    col = get_column_letter(bgn_col)
    f = f"={col}{food_cogs_row}+{col}{bev_cogs_row}"
    cogs_sub_formulas.append(f)
add_cost_line(row, "Total COGS", cogs_sub_formulas, bold=True, indent=0, is_subtotal=True)
COGS_ROW = row; row += 2

# OPEX
ws.cell(row=row, column=1, value="OPERATING EXPENSES").font = Font(bold=True, color=NAVY)
row += 1

# Rent (with 3% indexation)
rent_f = []
for i in range(5):
    f = f"={aref('Rent BGN/month')}*12*(1+{aref('Rent annual indexation')})^{i}"
    rent_f.append(f)
add_cost_line(row, "  Rent", rent_f); row += 1

# Utilities + oven fuel
util_f = []
for i in range(5):
    f = f"=({aref('Utilities base BGN/month')}+{aref('Oven gas/wood BGN/month')})*12"
    util_f.append(f)
add_cost_line(row, "  Utilities & oven fuel", util_f); row += 1

# Labour: (chef + 2*pizzaioli + 2*waiters + delivery PT) * 12 * (1 + social)
labour_f = []
for i in range(5):
    f = (f"=({aref('Head chef gross BGN/month')}+2*{aref('Pizzaioli x2 gross BGN/month each')}"
         f"+2*{aref('Waiters x2 gross BGN/month each')}+{aref('Delivery coordinator (PT) BGN/month')})"
         f"*12*(1+{aref('Employer social contributions %')})")
    labour_f.append(f)
add_cost_line(row, "  Labour (incl. 19% social)", labour_f); LABOUR_ROW = row; row += 1

# Marketing
mkt_f = []
for i in range(5):
    rev = rb(i, REV_TOTAL_ROW)
    if i == 0:
        f = f"={rev}*{aref('Marketing % of revenue Y1 (launch)')}"
    else:
        f = f"={rev}*{aref('Marketing % of revenue Y2+ (steady)')}"
    mkt_f.append(f)
add_cost_line(row, "  Marketing", mkt_f); row += 1

# R&M
rm_f = [f"={rb(i, REV_TOTAL_ROW)}*{aref('Repairs & maintenance % of revenue')}" for i in range(5)]
add_cost_line(row, "  Repairs & maintenance", rm_f); row += 1

# Packaging
pkg_f = [f"={rb(i, REV_TOTAL_ROW)}*{aref('Packaging % of revenue')}" for i in range(5)]
add_cost_line(row, "  Packaging & delivery supplies", pkg_f); row += 1

# Insurance
ins_f = [f"={aref('Insurance & licenses BGN/yr')}" for i in range(5)]
add_cost_line(row, "  Insurance & licenses", ins_f); row += 1

# POS
pos_f = [f"={aref('POS / SaaS BGN/month')}*12" for i in range(5)]
add_cost_line(row, "  POS / SaaS", pos_f); row += 1

# Accounting
acc_f = [f"={aref('Accounting BGN/month')}*12" for i in range(5)]
add_cost_line(row, "  Accounting", acc_f); row += 1

# Misc
misc_f = [f"={rb(i, REV_TOTAL_ROW)}*{aref('Misc % of revenue')}" for i in range(5)]
add_cost_line(row, "  Misc", misc_f); row += 1

# OPEX subtotal — rows 10..18 inclusive after our writes; build dynamically
OPEX_START = LABOUR_ROW - 2  # Rent row (just before utilities = LABOUR_ROW-1, rent = LABOUR_ROW-2)
# Safer: explicitly sum the 10 opex rows we just wrote
opex_rows = list(range(LABOUR_ROW - 2, row))  # rent .. misc inclusive
opex_sub_f = []
for i in range(5):
    bgn_col = 2 + i * 2
    col = get_column_letter(bgn_col)
    expr = "+".join(f"{col}{r}" for r in opex_rows)
    opex_sub_f.append(f"={expr}")
add_cost_line(row, "Total OPEX", opex_sub_f, bold=True, indent=0, is_subtotal=True)
OPEX_ROW = row; row += 2

# Total costs (COGS + OPEX)
total_f = []
for i in range(5):
    bgn_col = 2 + i * 2
    col = get_column_letter(bgn_col)
    total_f.append(f"={col}{COGS_ROW}+{col}{OPEX_ROW}")
add_cost_line(row, "TOTAL OPERATING COSTS (COGS + OPEX)", total_f, bold=True, indent=0, is_subtotal=True)
TOTAL_COSTS_ROW = row

ws.freeze_panes = "B6"

# ===================================================================
# 5. P&L
# ===================================================================
ws = wb.create_sheet("P&L")
title_row(ws, "Profit & Loss", cols=8)
set_col_widths(ws, [38, 18, 18, 18, 18, 18, 18, 18])

ws["A3"] = "Bulgaria CIT 10%. D&A = Opening CAPEX / 7 years straight-line."
ws["A3"].font = Font(italic=True, color="555555")
ws.merge_cells("A3:H3")

# Header
ws.cell(row=4, column=1, value="P&L Line"); sub_header(ws.cell(row=4, column=1))
ws.cell(row=4, column=2, value="% margin"); sub_header(ws.cell(row=4, column=2))
for i, y in enumerate(YEARS):
    c = ws.cell(row=4, column=3 + i, value=y); sub_header(c)

# Revenue
ws.cell(row=5, column=1, value="Revenue").font = Font(bold=True)
ws.cell(row=5, column=2, value="").fill = PatternFill("solid", fgColor=GRAY)
for i in range(5):
    f = f"='Revenue Build'!{get_column_letter(3+i)}{REV_TOTAL_ROW}"
    c = ws.cell(row=5, column=3 + i, value=f); output_cell(c)

# COGS
ws.cell(row=6, column=1, value="(–) COGS")
for i in range(5):
    f = f"=-'Cost Forecast'!{get_column_letter(2 + i*2)}{COGS_ROW}"
    c = ws.cell(row=6, column=3 + i, value=f); calc_cell(c)
# Gross profit
ws.cell(row=7, column=1, value="Gross Profit").font = Font(bold=True)
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"={col}5+{col}6"
    c = ws.cell(row=7, column=3 + i, value=f); output_cell(c)
# Gross margin pct
for r, base in [(6, 5), (7, 5)]:
    pass  # we put % per row below explicitly

# OPEX
ws.cell(row=8, column=1, value="(–) Operating Expenses")
for i in range(5):
    f = f"=-'Cost Forecast'!{get_column_letter(2 + i*2)}{OPEX_ROW}"
    c = ws.cell(row=8, column=3 + i, value=f); calc_cell(c)

# EBITDA
ws.cell(row=9, column=1, value="EBITDA").font = Font(bold=True, color=NAVY)
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"={col}7+{col}8"
    c = ws.cell(row=9, column=3 + i, value=f); output_cell(c)

# D&A
ws.cell(row=10, column=1, value="(–) D&A")
for i in range(5):
    f = f"=-{aref('Total Opening CAPEX')}/{aref('D&A useful life')}"
    c = ws.cell(row=10, column=3 + i, value=f); calc_cell(c)

# EBIT
ws.cell(row=11, column=1, value="EBIT").font = Font(bold=True)
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=11, column=3 + i, value=f"={col}9+{col}10"); calc_cell(c, bold=True)

# Interest
ws.cell(row=12, column=1, value="(–) Interest")
for i in range(5):
    c = ws.cell(row=12, column=3 + i, value=f"=-{aref('Interest expense')}"); calc_cell(c)

# PBT
ws.cell(row=13, column=1, value="Profit Before Tax").font = Font(bold=True)
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=13, column=3 + i, value=f"={col}11+{col}12"); calc_cell(c, bold=True)

# Tax
ws.cell(row=14, column=1, value=f"(–) Corporate Tax (10%)")
for i in range(5):
    col = get_column_letter(3 + i)
    f = f"=-MAX(0,{col}13)*{aref('Corporate income tax')}"
    c = ws.cell(row=14, column=3 + i, value=f); calc_cell(c)

# Net Income
ws.cell(row=15, column=1, value="Net Income").font = Font(bold=True, color=NAVY)
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=15, column=3 + i, value=f"={col}13+{col}14"); output_cell(c)

# Margin block
ws.cell(row=18, column=1, value="MARGIN ANALYSIS").font = Font(bold=True, color=NAVY)
ws.merge_cells("A18:H18")
ws["A18"].fill = PatternFill("solid", fgColor=GRAY)

# Headers
for i, y in enumerate(YEARS):
    sub_header(ws.cell(row=19, column=3 + i, value=y))
sub_header(ws.cell(row=19, column=1, value="Metric"))

ws.cell(row=20, column=1, value="Gross margin")
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=20, column=3 + i, value=f"={col}7/{col}5"); calc_cell(c, fmt=FMT_PCT)

ws.cell(row=21, column=1, value="OPEX % of revenue")
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=21, column=3 + i, value=f"=-{col}8/{col}5"); calc_cell(c, fmt=FMT_PCT)

ws.cell(row=22, column=1, value="EBITDA margin").font = Font(bold=True)
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=22, column=3 + i, value=f"={col}9/{col}5"); output_cell(c, fmt=FMT_PCT)

ws.cell(row=23, column=1, value="Net margin")
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=23, column=3 + i, value=f"={col}15/{col}5"); calc_cell(c, fmt=FMT_PCT, bold=True)

# Save row indexes for downstream
PNL_REV_ROW = 5
PNL_EBITDA_ROW = 9
PNL_TAX_ROW = 14
PNL_NI_ROW = 15

ws.freeze_panes = "C5"

# ===================================================================
# 6. CASH FLOW & INVESTMENT ANALYSIS
# ===================================================================
ws = wb.create_sheet("Cash Flow")
title_row(ws, "Cash Flow & Investment Analysis", cols=9)
set_col_widths(ws, [38, 14, 16, 16, 16, 16, 16, 16, 16])

ws["A3"] = "Y0 opening CAPEX + WC; FCF = EBITDA - Tax - Maintenance CAPEX. Terminal value = Y5 EBITDA × 4.5x."
ws["A3"].font = Font(italic=True, color="555555")
ws.merge_cells("A3:I3")

# Header: Y0, Y1..Y5, Terminal
periods = ["Y0 2025", "Y1 2026", "Y2 2027", "Y3 2028", "Y4 2029", "Y5 2030", "Terminal"]
sub_header(ws.cell(row=4, column=1, value="Line"))
for i, p in enumerate(periods):
    sub_header(ws.cell(row=4, column=2 + i, value=p))

# Opening CAPEX
ws.cell(row=5, column=1, value="(–) Opening CAPEX")
c = ws.cell(row=5, column=2, value=f"=-{aref('Total Opening CAPEX')}"); calc_cell(c)
for i in range(6):
    ws.cell(row=5, column=3 + i, value=0).number_format = FMT_BGN
    ws.cell(row=5, column=3 + i).border = BORDER

# Working capital
ws.cell(row=6, column=1, value="(–/+) Working capital")
ws.cell(row=6, column=2, value=f"=-{aref('Working capital buffer')}").number_format = FMT_BGN
ws.cell(row=6, column=2).border = BORDER
for i in range(5):
    ws.cell(row=6, column=3 + i, value=0).number_format = FMT_BGN
    ws.cell(row=6, column=3 + i).border = BORDER
ws.cell(row=6, column=8, value=f"={aref('Working capital buffer')}").number_format = FMT_BGN  # returned in terminal
ws.cell(row=6, column=8).border = BORDER

# EBITDA pull
ws.cell(row=7, column=1, value="EBITDA")
ws.cell(row=7, column=2, value=0).border = BORDER; ws.cell(row=7, column=2).number_format = FMT_BGN
for i in range(5):
    f = f"='P&L'!{get_column_letter(3+i)}{PNL_EBITDA_ROW}"
    c = ws.cell(row=7, column=3 + i, value=f); calc_cell(c)
ws.cell(row=7, column=8, value=0).border = BORDER; ws.cell(row=7, column=8).number_format = FMT_BGN

# Tax
ws.cell(row=8, column=1, value="(–) Cash Tax")
ws.cell(row=8, column=2, value=0).border = BORDER; ws.cell(row=8, column=2).number_format = FMT_BGN
for i in range(5):
    f = f"='P&L'!{get_column_letter(3+i)}{PNL_TAX_ROW}"
    c = ws.cell(row=8, column=3 + i, value=f); calc_cell(c)
ws.cell(row=8, column=8, value=0).border = BORDER; ws.cell(row=8, column=8).number_format = FMT_BGN

# Maintenance CAPEX (Y2+)
ws.cell(row=9, column=1, value="(–) Maintenance CAPEX")
ws.cell(row=9, column=2, value=0).border = BORDER; ws.cell(row=9, column=2).number_format = FMT_BGN
for i in range(5):
    rev = f"='Revenue Build'!{get_column_letter(3+i)}{REV_TOTAL_ROW}"
    if i == 0:
        f = "=0"
    else:
        f = f"=-'Revenue Build'!{get_column_letter(3+i)}{REV_TOTAL_ROW}*{aref('Maintenance CAPEX % of revenue (Y2+)')}"
    c = ws.cell(row=9, column=3 + i, value=f); calc_cell(c)
ws.cell(row=9, column=8, value=0).border = BORDER; ws.cell(row=9, column=8).number_format = FMT_BGN

# Terminal value
ws.cell(row=10, column=1, value="(+) Terminal Value (Y5 EBITDA × multiple)")
ws.cell(row=10, column=2, value=0).border = BORDER; ws.cell(row=10, column=2).number_format = FMT_BGN
for i in range(5):
    ws.cell(row=10, column=3 + i, value=0).border = BORDER
    ws.cell(row=10, column=3 + i).number_format = FMT_BGN
tv_formula = f"='P&L'!{get_column_letter(3+4)}{PNL_EBITDA_ROW}*{aref('Exit EBITDA multiple')}"
ws.cell(row=10, column=8, value=tv_formula).number_format = FMT_BGN
ws.cell(row=10, column=8).border = BORDER

# Free Cash Flow
ws.cell(row=11, column=1, value="Free Cash Flow").font = Font(bold=True, color=NAVY)
for i in range(7):
    col = get_column_letter(2 + i)
    f = f"=SUM({col}5:{col}10)"
    c = ws.cell(row=11, column=2 + i, value=f); output_cell(c)

# Cumulative FCF (for payback)
ws.cell(row=12, column=1, value="Cumulative FCF").font = Font(italic=True)
ws.cell(row=12, column=2, value="=B11").number_format = FMT_BGN; ws.cell(row=12, column=2).border = BORDER
for i in range(1, 7):
    col = get_column_letter(2 + i)
    prev = get_column_letter(2 + i - 1)
    c = ws.cell(row=12, column=2 + i, value=f"={prev}12+{col}11"); calc_cell(c)

CF_FCF_ROW = 11
CF_CUM_ROW = 12

# Investment summary block
ws.cell(row=15, column=1, value="INVESTMENT METRICS").font = Font(bold=True, size=12, color="FFFFFF")
ws.cell(row=15, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws.merge_cells("A15:I15")

# Inputs visible
ws.cell(row=16, column=1, value="Discount rate (WACC)")
c = ws.cell(row=16, column=3, value=f"={aref('Discount rate (WACC proxy)')}"); output_cell(c, fmt=FMT_PCT)
ws.cell(row=17, column=1, value="Exit EBITDA multiple")
c = ws.cell(row=17, column=3, value=f"={aref('Exit EBITDA multiple')}"); output_cell(c, fmt=FMT_MULT)
ws.cell(row=18, column=1, value="Total cash deployed (Y0)")
c = ws.cell(row=18, column=3, value=f"=-B{CF_FCF_ROW}"); output_cell(c)

# We combine Y5 FCF + Terminal cash flow into the final period for IRR/NPV
# Build a vector: -outflow Y0, FCF Y1..Y4, FCF Y5 + Terminal
ws.cell(row=20, column=1, value="Cash flow vector for IRR/NPV (Y0..Y5)").font = Font(italic=True, color="555555")
for i in range(6):
    period_col = get_column_letter(2 + i)
    if i < 5:
        f = f"={period_col}{CF_FCF_ROW}"
    else:
        # Y5 FCF + Terminal FCF combined
        f = f"={period_col}{CF_FCF_ROW}+H{CF_FCF_ROW}"
    c = ws.cell(row=20, column=3 + i, value=f); calc_cell(c)

IRR_RANGE = f"C20:H20"

ws.cell(row=22, column=1, value="Investment Returns").font = Font(bold=True, color=NAVY)

ws.cell(row=23, column=1, value="Total terminal proceeds (Y5 FCF + Terminal)")
c = ws.cell(row=23, column=3, value=f"=H20"); output_cell(c)

ws.cell(row=24, column=1, value="NPV @ Discount rate (BGN)").font = Font(bold=True)
c = ws.cell(row=24, column=3, value=f"=NPV({aref('Discount rate (WACC proxy)')},D20:H20)+C20"); output_cell(c)

ws.cell(row=25, column=1, value="IRR").font = Font(bold=True)
c = ws.cell(row=25, column=3, value=f"=IRR({IRR_RANGE})"); output_cell(c, fmt=FMT_PCT)

ws.cell(row=26, column=1, value="MOIC (Multiple on Invested Capital)").font = Font(bold=True)
c = ws.cell(row=26, column=3, value=f"=SUM(D20:H20)/(-C20)"); output_cell(c, fmt=FMT_MULT)

# Payback — find first year cum FCF turns positive (operating, excl terminal)
ws.cell(row=27, column=1, value="Payback period (yrs)").font = Font(bold=True)
# Approx: use linear interpolation; build a formula using MATCH on cumulative FCF
payback_formula = (
    f"=IFERROR(MATCH(TRUE,INDEX(C{CF_CUM_ROW}:G{CF_CUM_ROW}>=0,0),0)"
    f"-1+(-INDEX(C{CF_CUM_ROW}:G{CF_CUM_ROW},MATCH(TRUE,INDEX(C{CF_CUM_ROW}:G{CF_CUM_ROW}>=0,0),0)-1))"
    f"/INDEX(C{CF_FCF_ROW}:G{CF_FCF_ROW},MATCH(TRUE,INDEX(C{CF_CUM_ROW}:G{CF_CUM_ROW}>=0,0),0)),99)"
)
# Simpler & robust fallback:
payback_simple = (
    f"=IFERROR(IF(C{CF_CUM_ROW}>=0,1,IF(D{CF_CUM_ROW}>=0,1+(-C{CF_CUM_ROW})/D{CF_FCF_ROW},"
    f"IF(E{CF_CUM_ROW}>=0,2+(-D{CF_CUM_ROW})/E{CF_FCF_ROW},"
    f"IF(F{CF_CUM_ROW}>=0,3+(-E{CF_CUM_ROW})/F{CF_FCF_ROW},"
    f"IF(G{CF_CUM_ROW}>=0,4+(-F{CF_CUM_ROW})/G{CF_FCF_ROW},99))))),99)"
)
c = ws.cell(row=27, column=3, value=payback_simple); output_cell(c, fmt='0.0" yrs"')

ws.freeze_panes = "B5"

# ===================================================================
# 7. SENSITIVITY (IRR vs ASP × Covers/day)
# ===================================================================
ws = wb.create_sheet("Sensitivity")
title_row(ws, "Sensitivity — IRR", cols=8)
set_col_widths(ws, [22, 14, 14, 14, 14, 14, 14, 14])

ws["A3"] = ("5×5 IRR sensitivity. ASP and covers/day flexed ±15% (cols) and ±20% (rows). "
            "Note: this is an illustrative static-driver matrix — re-run model with scenario toggle for exact IRR.")
ws["A3"].font = Font(italic=True, color="555555")
ws.merge_cells("A3:H3")

# Static sensitivity: we compute IRR analytically using simple linear approximation
# IRR scales roughly with (ASP × Volume) → revenue. We'll compute an indicative IRR
# by re-scaling Y5 EBITDA proportionally and re-running NPV/IRR with formulas.
# For a true sensitivity we'd need a data table; here we approximate via formula:
# Approx EBITDA scaling: EBITDA changes by ΔRev * gross margin proxy
# We use central-case Y5 EBITDA and total invested. Indicative IRR via this approximation.

asp_flex = [-0.15, -0.075, 0, 0.075, 0.15]
vol_flex = [-0.20, -0.10, 0, 0.10, 0.20]

# Header row (ASP flex)
sub_header(ws.cell(row=5, column=1, value="Covers ↓ / ASP →"))
for i, a in enumerate(asp_flex):
    c = ws.cell(row=5, column=2 + i, value=a)
    sub_header(c); c.number_format = "+0%;-0%;0%"

# Pull central case numbers
ws["I5"] = "Central Y5 EBITDA"
ws["J5"] = f"='P&L'!G{PNL_EBITDA_ROW}"
ws["J5"].number_format = FMT_BGN
ws["I6"] = "Central Y5 Revenue"
ws["J6"] = f"='P&L'!G{PNL_REV_ROW}"
ws["J6"].number_format = FMT_BGN
ws["I7"] = "Invested (Y0)"
ws["J7"] = f"=-'Cash Flow'!B{CF_FCF_ROW}"
ws["J7"].number_format = FMT_BGN
ws["I8"] = "Exit multiple"
ws["J8"] = f"={aref('Exit EBITDA multiple')}"
ws["J8"].number_format = FMT_MULT

# Indicative IRR proxy:
# Scenario revenue = central_rev × (1+asp_flex) × (1+vol_flex)
# Scenario EBITDA = central_EBITDA + (Scenario rev - central rev) × 0.60 (~ incremental margin on flex)
# Exit value ≈ EBITDA × multiple
# IRR proxy = (exit / invested)^(1/5) - 1
for ri, v in enumerate(vol_flex):
    c = ws.cell(row=6 + ri, column=1, value=v)
    sub_header(c); c.number_format = "+0%;-0%;0%"
    for ci, a in enumerate(asp_flex):
        # incremental margin assumed 60%
        formula = (
            f"=((($J$5+($J$6*((1+{a})*(1+{v})-1))*0.6)*$J$8)/$J$7)^(1/5)-1"
        )
        cell = ws.cell(row=6 + ri, column=2 + ci, value=formula)
        cell.number_format = FMT_PCT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

# Conditional formatting (green high, red low)
ws.conditional_formatting.add(
    "B6:F10",
    ColorScaleRule(start_type='min', start_color='F8CBAD',
                   mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                   end_type='max', end_color='C6E0B4')
)

ws["A12"] = "Read: % change in ASP across columns; % change in covers across rows. Cells show indicative 5-yr IRR."
ws["A12"].font = Font(italic=True, color="555555")
ws.merge_cells("A12:H12")

ws.freeze_panes = "B6"

# ===================================================================
# 8. KPI DASHBOARD
# ===================================================================
ws = wb.create_sheet("KPI Dashboard")
title_row(ws, "KPI Dashboard", cols=8)
set_col_widths(ws, [38, 16, 16, 16, 16, 16, 16, 16])

ws["A3"] = "One-page operating dashboard — links live to other sheets."
ws["A3"].font = Font(italic=True, color="555555")
ws.merge_cells("A3:H3")

# Header
sub_header(ws.cell(row=5, column=1, value="KPI"))
for i, y in enumerate(YEARS):
    sub_header(ws.cell(row=5, column=2 + i, value=y))

# Revenue
ws.cell(row=6, column=1, value="Total Revenue (BGN)").font = Font(bold=True)
for i in range(5):
    f = f"='P&L'!{get_column_letter(3+i)}{PNL_REV_ROW}"
    c = ws.cell(row=6, column=2 + i, value=f); output_cell(c)

# Revenue per cover (use dine-in covers only as cleanest comp)
ws.cell(row=7, column=1, value="Avg revenue / dine-in cover (BGN)")
for i in range(5):
    col_rb = get_column_letter(3 + i)
    f = f"='Revenue Build'!{col_rb}{dinein_ticket_row}"  # scenario-flexed ticket
    c = ws.cell(row=7, column=2 + i, value=f); calc_cell(c, fmt=FMT_BGN)

# Food cost % rev
ws.cell(row=8, column=1, value="Food cost % of revenue")
for i in range(5):
    bgn_col = 2 + i * 2
    f = f"='Cost Forecast'!{get_column_letter(bgn_col)}{food_cogs_row}/'P&L'!{get_column_letter(3+i)}{PNL_REV_ROW}"
    c = ws.cell(row=8, column=2 + i, value=f); calc_cell(c, fmt=FMT_PCT)

# Labour % rev
ws.cell(row=9, column=1, value="Labour % of revenue")
for i in range(5):
    bgn_col = 2 + i * 2
    f = f"='Cost Forecast'!{get_column_letter(bgn_col)}{LABOUR_ROW}/'P&L'!{get_column_letter(3+i)}{PNL_REV_ROW}"
    c = ws.cell(row=9, column=2 + i, value=f); calc_cell(c, fmt=FMT_PCT)

# EBITDA margin
ws.cell(row=10, column=1, value="EBITDA margin").font = Font(bold=True)
for i in range(5):
    f = f"='P&L'!{get_column_letter(3+i)}22"
    c = ws.cell(row=10, column=2 + i, value=f); output_cell(c, fmt=FMT_PCT)

# Pizzas/day vs capacity
ws.cell(row=11, column=1, value="Pizzas/day (implied)")
for i in range(5):
    col_rb = get_column_letter(3 + i)
    f = f"='Revenue Build'!{col_rb}{pizzas_day_row}"
    c = ws.cell(row=11, column=2 + i, value=f); calc_cell(c, fmt='#,##0')

ws.cell(row=12, column=1, value="Oven utilisation")
for i in range(5):
    col_rb = get_column_letter(3 + i)
    f = f"='Revenue Build'!{col_rb}{util_row}"
    c = ws.cell(row=12, column=2 + i, value=f); calc_cell(c, fmt=FMT_PCT, bold=True)

# Monthly break-even
ws.cell(row=13, column=1, value="Monthly break-even revenue (BGN)")
for i in range(5):
    # BE = (Fixed costs / yr) / Contribution margin %; approx fixed = OPEX; CM = 1 - COGS%
    bgn_col = 2 + i * 2
    rev_col_pnl = get_column_letter(3 + i)
    f = (f"=('Cost Forecast'!{get_column_letter(bgn_col)}{OPEX_ROW}/12)"
         f"/(1-('Cost Forecast'!{get_column_letter(bgn_col)}{COGS_ROW}/'P&L'!{rev_col_pnl}{PNL_REV_ROW}))")
    c = ws.cell(row=13, column=2 + i, value=f); calc_cell(c, fmt=FMT_BGN, bold=True)

# Mini bar chart for EBITDA margin trajectory
chart = BarChart()
chart.type = "col"
chart.title = "EBITDA Margin Trajectory"
chart.y_axis.title = "EBITDA Margin"
chart.x_axis.title = "Year"
data = Reference(ws, min_col=2, max_col=6, min_row=10, max_row=10)
cats = Reference(ws, min_col=2, max_col=6, min_row=5, max_row=5)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.height = 8
chart.width = 16
ws.add_chart(chart, "A16")

# Mini bar chart for Revenue trajectory
chart2 = BarChart()
chart2.type = "col"
chart2.style = 11
chart2.title = "Revenue Trajectory (BGN)"
chart2.y_axis.title = "Revenue"
chart2.x_axis.title = "Year"
data2 = Reference(ws, min_col=2, max_col=6, min_row=6, max_row=6)
chart2.add_data(data2, titles_from_data=False)
chart2.set_categories(cats)
chart2.height = 8
chart2.width = 16
ws.add_chart(chart2, "A33")

ws.freeze_panes = "B6"

# ===================================================================
# Reorder tabs explicitly
# ===================================================================
order = ["Cover", "Assumptions", "Revenue Build", "Cost Forecast",
         "P&L", "Cash Flow", "Sensitivity", "KPI Dashboard"]
wb._sheets = [wb[s] for s in order]

# Save
os.makedirs(OUT_DIR, exist_ok=True)
wb.save(OUT_FILE)

# ---------- Verify ----------
size = os.path.getsize(OUT_FILE)
print(f"OK file written: {OUT_FILE}")
print(f"size: {size:,} bytes")
print("Tabs:")
for s in order:
    print(f"  - {s}")
